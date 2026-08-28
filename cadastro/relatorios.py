"""Geração de relatórios (Fase 5): planilhas Excel e PDFs.

- ``planilha_response`` — exporta cabeçalhos + linhas para .xlsx (openpyxl).
- ``ficha_pdf`` — ficha individual de um Núcleo Familiar (reportlab).
- ``classificacao_pdf`` — lista de classificação (reportlab).
"""

from __future__ import annotations

from io import BytesIO

from django.contrib.staticfiles import finders
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Image,
    KeepTogether,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_AZUL = colors.HexColor("#1f4e79")

# Caracteres que o Excel/LibreOffice interpretam como início de fórmula.
_GATILHOS_FORMULA = ("=", "+", "-", "@", "\t", "\r", "\n")


def _neutralizar(valor):
    """Evita injeção de fórmula (CSV injection) em células de texto.

    Dados de terceiros (ex.: nome do requerente) podem começar com '=' e virar
    fórmula ativa na planilha. Prefixa um apóstrofo para forçar texto.
    """
    if isinstance(valor, str) and valor.startswith(_GATILHOS_FORMULA):
        return "'" + valor
    return valor


def planilha_response(filename: str, cabecalhos: list[str], linhas) -> HttpResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"
    ws.append(cabecalhos)
    for celula in ws[1]:
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor="1F4E79")
    for linha in linhas:
        ws.append([_neutralizar(v) for v in linha])
    for i, _ in enumerate(cabecalhos, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 22
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.getvalue(), content_type=XLSX_MIME)
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def _pdf_response(nome: str, elementos) -> HttpResponse:
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4, topMargin=18 * mm, bottomMargin=18 * mm,
        leftMargin=16 * mm, rightMargin=16 * mm, title=nome,
    )
    doc.build(elementos)
    buf.seek(0)
    resp = HttpResponse(buf.getvalue(), content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="{nome}"'
    return resp


def _estilo_tabela(cabecalho=True):
    cmds = [
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d6dce3")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f6f8")]),
    ]
    if cabecalho:
        cmds += [
            ("BACKGROUND", (0, 0), (-1, 0), _AZUL),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ]
    return TableStyle(cmds)


def _checklist_flowables(inscricao, estilos, titulo="Checklist da documentação (item 4)", fonte=8.0):
    """Tabela do checklist da documentação exigida (Apresentado/Falta)."""
    from . import documentos as _docs

    itens = _docs.exigidos(inscricao)
    flows = [Paragraph(titulo, estilos["Heading3"])]
    linhas = [["Item", "Documento", "Situação"]]
    for ex in itens:
        rot = ex.rotulo + (f" ({ex.detalhe})" if ex.detalhe else "")
        linhas.append([ex.codigo, rot, "Apresentado" if ex.ok else "FALTA"])
    t = Table(linhas, colWidths=[15 * mm, 129 * mm, 30 * mm], repeatRows=1)
    cmds = [
        ("FONTSIZE", (0, 0), (-1, -1), fonte),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d6dce3")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BACKGROUND", (0, 0), (-1, 0), _AZUL),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]
    for i, ex in enumerate(itens, start=1):
        cor = colors.HexColor("#137333") if ex.ok else colors.HexColor("#b3261e")
        cmds.append(("TEXTCOLOR", (2, i), (2, i), cor))
        cmds.append(("FONTNAME", (2, i), (2, i), "Helvetica-Bold"))
    t.setStyle(TableStyle(cmds))
    flows.append(t)
    return flows


def ficha_pdf(inscricao) -> HttpResponse:
    estilos = getSampleStyleSheet()
    h = estilos["Heading2"]
    h.textColor = _AZUL
    normal = estilos["BodyText"]
    e = []

    e.append(Paragraph("Ficha do Núcleo Familiar — MCMV / Barão de Cocais", estilos["Title"]))
    e.append(Paragraph(
        f"Inscrição <b>{inscricao.numero_inscricao}</b> — situação: "
        f"{inscricao.get_status_display()}", normal))
    e.append(Spacer(1, 6))

    req = inscricao.requerente
    dados = [
        ["Requerente", req.nome],
        ["CPF", req.cpf_fmt],
        ["Nascimento", str(req.data_nascimento)],
        ["Contato", f"{inscricao.telefone}  {inscricao.email}"],
        ["Endereço", f"{inscricao.endereco} {inscricao.numero} {inscricao.bairro} {inscricao.cep}"],
    ]
    t = Table(dados, colWidths=[35 * mm, 140 * mm])
    t.setStyle(_estilo_tabela(cabecalho=False))
    e.append(t)

    e.append(Paragraph("Composição do núcleo", h))
    linhas = [["Nome", "Parentesco", "Nascimento", "Arrimo", "PcD"]]
    for m in inscricao.membros.select_related("pessoa"):
        linhas.append([
            m.pessoa.nome, m.get_parentesco_display(), str(m.pessoa.data_nascimento),
            "Sim" if m.arrimo else "—", "Sim" if m.pessoa.pcd else "—",
        ])
    t = Table(linhas, colWidths=[60 * mm, 45 * mm, 30 * mm, 20 * mm, 20 * mm])
    t.setStyle(_estilo_tabela())
    e.append(t)

    e.append(Paragraph("Critérios Legais", h))
    linhas = [["Inciso", "Atendido", "Pontos"]]
    for c in inscricao.criterios_legais.all():
        linhas.append([c.get_inciso_display(), "Sim" if c.atendido else "Não", str(c.pontos)])
    t = Table(linhas, colWidths=[120 * mm, 30 * mm, 25 * mm])
    t.setStyle(_estilo_tabela())
    e.append(t)

    cc = getattr(inscricao, "criterio_complementar", None)
    if cc:
        e.append(Paragraph("Critérios Complementares", h))
        linhas = [
            ["Item", "Apurado", "Pontos"],
            ["Renda per capita", f"R$ {cc.renda_per_capita}", str(cc.pontos_renda)],
            ["Comprometimento aluguel", f"{cc.percentual or '—'}%", str(cc.pontos_aluguel)],
        ]
        t = Table(linhas, colWidths=[100 * mm, 45 * mm, 30 * mm])
        t.setStyle(_estilo_tabela())
        e.append(t)

    e.append(Spacer(1, 8))
    e.append(Paragraph(
        f"<b>CL = {inscricao.pontos_legais or 0} &nbsp; "
        f"CC = {inscricao.pontos_complementares or 0} &nbsp; "
        f"→ &nbsp; P = {inscricao.pontuacao_total or 0} pontos</b> (máx. 190)",
        estilos["Heading3"],
    ))

    # Checklist da documentação exigida (item 4 do edital).
    e.append(Spacer(1, 8))
    e.extend(_checklist_flowables(inscricao, estilos))

    # Pendências de cadastro (campos esperados em branco).
    from .wizard import pendencias as _pendencias

    pend = _pendencias(inscricao)
    e.append(Paragraph("Pendências de cadastro", h))
    if pend:
        for item in pend:
            e.append(Paragraph(f"• {item}", normal))
    else:
        e.append(Paragraph("Nenhuma pendência — cadastro completo.", normal))

    return _pdf_response(f"ficha_{inscricao.numero_inscricao}.pdf", e)


def classificacao_pdf(itens) -> HttpResponse:
    estilos = getSampleStyleSheet()
    e = [Paragraph("Classificação Geral — MCMV / Barão de Cocais", estilos["Title"]),
         Spacer(1, 6)]
    linhas = [["Pos.", "Inscrição", "Requerente", "Pontos", "Filhos ≤12", "Idosos", "Empate"]]
    for c in itens:
        linhas.append([
            str(c.posicao or "—"), c.inscricao.numero_inscricao, c.inscricao.requerente.nome,
            str(c.pontuacao), str(c.dependentes_ate_12), str(c.idosos),
            "Sorteio" if c.empate_pendente_sorteio else "—",
        ])
    t = Table(linhas, colWidths=[14 * mm, 26 * mm, 60 * mm, 18 * mm, 22 * mm, 16 * mm, 20 * mm], repeatRows=1)
    t.setStyle(_estilo_tabela())
    e.append(t)
    return _pdf_response("classificacao.pdf", e)


def _nome_usuario(u) -> str:
    """Nome de exibição de um servidor (nome completo ou, na falta, o login)."""
    if not u:
        return ""
    return (u.get_full_name() or u.get_username()).strip()


def lista_ordem_cadastro_pdf(inscricoes) -> HttpResponse:
    """Relação geral das inscrições por ordem de cadastro (ordem de chegada).

    Lista simples, ordenada pela data/hora da inscrição (item 2.1 do edital),
    independentemente da pontuação — útil para conferência do protocolo.
    """
    from django.utils import timezone

    estilos = getSampleStyleSheet()
    e = [
        Paragraph("Relação de inscrições por ordem de cadastro", estilos["Title"]),
        Paragraph("MCMV — Barão de Cocais/MG · Edital 001/2026", estilos["Normal"]),
        Spacer(1, 8),
    ]
    linhas = [["#", "Nº", "Data/hora do cadastro", "Requerente", "CPF", "Situação"]]
    for ordem, i in enumerate(inscricoes, start=1):
        dt = timezone.localtime(i.data_inscricao).strftime("%d/%m/%Y %H:%M")
        linhas.append([
            str(ordem), i.numero_inscricao, dt,
            i.requerente.nome, i.requerente.cpf_fmt, i.get_status_display(),
        ])
    t = Table(linhas, colWidths=[10 * mm, 24 * mm, 34 * mm, 55 * mm, 32 * mm, 30 * mm], repeatRows=1)
    t.setStyle(_estilo_tabela())
    e.append(t)
    e.append(Spacer(1, 8))
    e.append(Paragraph(
        f"Total de inscrições: {len(linhas) - 1}. A ordem de cadastro não é "
        "critério de classificação (item 7.2 do edital).", estilos["Normal"]))
    return _pdf_response("inscricoes_ordem_cadastro.pdf", e)


def _recibo_via(inscricao, titulo_via: str, estilos) -> list:
    """Monta os elementos de UMA via do comprovante (compacta, meia folha)."""
    via_lbl = estilos["Normal"].clone("via_lbl")
    via_lbl.fontSize = 8
    via_lbl.textColor = colors.HexColor("#5a6672")
    via_lbl.alignment = 2  # direita

    cab = estilos["Normal"].clone("cab_recibo")
    cab.fontSize = 9
    cab.alignment = 1  # centro

    titulo = estilos["Heading2"].clone("tit_recibo")
    titulo.alignment = 1
    titulo.textColor = _AZUL

    corpo = estilos["Normal"].clone("corpo_recibo")
    corpo.fontSize = 8

    flows = [Paragraph(titulo_via, via_lbl)]

    # Cabeçalho compacto: logo pequeno + textos centrados.
    caminho_logo = finders.find("img/logo-prefeitura.jpg")
    if caminho_logo:
        img = Image(caminho_logo, width=16 * mm, height=16 * mm)
        img.hAlign = "CENTER"
        flows.append(img)
    flows.append(Paragraph("Prefeitura Municipal de Barão de Cocais/MG", cab))
    flows.append(Paragraph("Secretaria Municipal de Assistência Social", cab))
    flows.append(Spacer(1, 3))
    flows.append(Paragraph("COMPROVANTE DE INSCRIÇÃO", titulo))
    flows.append(Paragraph(
        "Programa Minha Casa, Minha Vida — Faixa 02 · Edital 001/2026", cab))
    flows.append(Spacer(1, 5))

    from django.utils import timezone

    data_fin = inscricao.data_finalizacao
    dados = [
        ["Protocolo", inscricao.protocolo or "—"],
        ["Nº da inscrição", inscricao.numero_inscricao],
        ["Data/hora", timezone.localtime(data_fin).strftime("%d/%m/%Y %H:%M") if data_fin else "—"],
        ["Requerente", inscricao.requerente.nome],
        ["CPF", inscricao.requerente.cpf_fmt],
        ["Integrantes do núcleo", str(inscricao.membros.count())],
        ["Situação", inscricao.get_status_display()],
    ]
    t = Table(dados, colWidths=[45 * mm, 129 * mm])
    est = _estilo_tabela(cabecalho=False)
    est.add("FONTSIZE", (0, 0), (-1, -1), 8)
    est.add("TOPPADDING", (0, 0), (-1, -1), 1.5)
    est.add("BOTTOMPADDING", (0, 0), (-1, -1), 1.5)
    t.setStyle(est)
    flows.append(t)

    # Resumo compacto da documentação (item 4) — o checklist detalhado vai na ficha.
    from . import documentos as _docs

    itens = _docs.exigidos(inscricao)
    apres = sum(1 for x in itens if x.ok)
    falt = [x.rotulo for x in itens if not x.ok]
    flows.append(Spacer(1, 4))
    flows.append(Paragraph(
        f"<b>Documentação (item 4):</b> {apres} de {len(itens)} itens apresentados.", corpo))
    if falt:
        txt = "; ".join(falt) if len(falt) <= 8 else f"{len(falt)} itens (ver ficha do cadastro)"
        flows.append(Paragraph(f"Pendentes: {txt}.", corpo))

    flows.append(Spacer(1, 4))
    flows.append(Paragraph(
        "Declaramos que a inscrição acima foi recebida nesta data. A inscrição e a "
        "eventual classificação <b>não geram direito</b> à contratação, financiamento "
        "ou aquisição de unidade, que dependem de análise posterior da instituição "
        "financeira, nos termos do Edital 001/2026.", corpo))

    flows.append(Spacer(1, 10))
    servidor = _nome_usuario(inscricao.finalizado_por)
    requerente = inscricao.requerente.nome
    assinaturas = [
        ["_______________________________", "_______________________________"],
        [servidor or " ", requerente or " "],
        ["Servidor responsável", "Requerente"],
    ]
    ta = Table(assinaturas, colWidths=[87 * mm, 87 * mm])
    ta.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("TOPPADDING", (0, 1), (-1, 1), 2),
        ("TEXTCOLOR", (0, 2), (-1, 2), colors.HexColor("#5a6672")),
    ]))
    flows.append(ta)
    return flows


def recibo_pdf(inscricao) -> HttpResponse:
    """Comprovante de inscrição em DUAS VIAS numa folha (Assistência e Requerente)."""
    estilos = getSampleStyleSheet()
    corte = estilos["Normal"].clone("corte")
    corte.fontSize = 7.5
    corte.textColor = colors.HexColor("#8a97a3")
    corte.alignment = 1

    e = []
    e.append(KeepTogether(_recibo_via(inscricao, "1ª VIA — ASSISTÊNCIA SOCIAL", estilos)))
    e.append(Spacer(1, 6))
    e.append(Paragraph("✂ - - - - - - - - - - - - - - - - - - - -  corte aqui  "
                       "- - - - - - - - - - - - - - - - - - - -", corte))
    e.append(Spacer(1, 6))
    e.append(KeepTogether(_recibo_via(inscricao, "2ª VIA — REQUERENTE", estilos)))

    # Comprovante: margens mais justas para caber as duas vias em uma folha.
    nome = f"recibo_{inscricao.protocolo or inscricao.numero_inscricao}.pdf"
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4, topMargin=12 * mm, bottomMargin=12 * mm,
        leftMargin=16 * mm, rightMargin=16 * mm, title=nome,
    )
    doc.build(e)
    buf.seek(0)
    resp = HttpResponse(buf.getvalue(), content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="{nome}"'
    return resp
