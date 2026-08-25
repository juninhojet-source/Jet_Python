"""Testes de integração da Fase 2 (Django): pontuação ponta a ponta, trilha de
auditoria e bloqueio pós-finalização.

Rodar: ``python manage.py test``
(Os testes do motor puro ficam em ``tests/`` e rodam com ``pytest``.)
"""

from datetime import date
from decimal import Decimal

from django.contrib.auth.models import User
from django.core.exceptions import ValidationError
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone

from auditoria.models import Auditoria

from . import requisitos, services
from .models import Inscricao, MembroNucleo, Pessoa, Renda

REF = date(2026, 9, 15)


def nasc(idade):
    return date(REF.year - idade, REF.month, REF.day)


def gera_cpf(base9: str) -> str:
    """Gera um CPF válido (com dígitos verificadores) a partir de 9 dígitos."""
    cpf = base9
    for _ in range(2):
        soma = sum(int(cpf[i]) * ((len(cpf) + 1) - i) for i in range(len(cpf)))
        cpf += str(((soma * 10) % 11) % 10)
    return cpf


class PontuacaoIntegracaoTest(TestCase):
    def _criar_inscricao_exemplo(self):
        req = Pessoa.objects.create(nome="João", cpf="111", data_nascimento=nasc(40), sexo="M")
        conj = Pessoa.objects.create(nome="Maria", cpf="222", data_nascimento=nasc(38), sexo="F")
        f1 = Pessoa.objects.create(nome="Pedro", cpf="333", data_nascimento=nasc(5), sexo="M")
        f2 = Pessoa.objects.create(nome="Ana", cpf="444", data_nascimento=nasc(8), sexo="F")

        insc = Inscricao.objects.create(
            requerente=req,
            data_referencia=REF,
            habitacao_precaria_ou_risco=True,
            matricula_comprovada=True,
            aluguel_mes_1=Decimal("1000"),
            aluguel_mes_2=Decimal("1100"),
            aluguel_mes_3=Decimal("1000"),
            status=Inscricao.Status.APTO,
        )
        m_req = MembroNucleo.objects.create(
            inscricao=insc, pessoa=req, parentesco="REQUERENTE", arrimo=True
        )
        Renda.objects.create(membro=m_req, tipo="FORMAL", valor=Decimal("3000"))
        MembroNucleo.objects.create(inscricao=insc, pessoa=conj, parentesco="COMPANHEIRO")
        MembroNucleo.objects.create(inscricao=insc, pessoa=f1, parentesco="FILHO")
        MembroNucleo.objects.create(inscricao=insc, pessoa=f2, parentesco="FILHO")
        return insc

    def test_exemplo_141_pontos_ponta_a_ponta(self):
        insc = self._criar_inscricao_exemplo()
        r = services.calcular_e_salvar(insc)

        insc.refresh_from_db()
        self.assertEqual(insc.pontos_legais, 120)
        self.assertEqual(insc.pontos_complementares, 21)
        self.assertEqual(insc.pontuacao_total, 141)
        self.assertEqual(r.pontuacao_total, 141)

        # Critérios legais persistidos
        self.assertEqual(insc.criterios_legais.get(inciso="CL_III").pontos, 0)
        self.assertEqual(insc.criterios_legais.get(inciso="CL_IV").pontos, 40)
        # Classificação criada com chaves de desempate
        self.assertEqual(insc.classificacao.dependentes_ate_12, 2)
        self.assertEqual(insc.classificacao.idosos, 0)

    def test_numero_inscricao_gerado(self):
        insc = self._criar_inscricao_exemplo()
        self.assertTrue(insc.numero_inscricao)
        self.assertEqual(insc.numero_inscricao, f"{insc.pk:06d}")


class AuditoriaTest(TestCase):
    def test_criacao_gera_registro(self):
        Pessoa.objects.create(nome="Zé", cpf="999", data_nascimento=nasc(30))
        self.assertTrue(
            Auditoria.objects.filter(
                tabela="cadastro.Pessoa", operacao=Auditoria.Operacao.CRIACAO
            ).exists()
        )

    def test_alteracao_registra_campo(self):
        p = Pessoa.objects.create(nome="Zé", cpf="998", data_nascimento=nasc(30))
        p.nome = "José"
        p.save()
        reg = Auditoria.objects.get(
            tabela="cadastro.Pessoa",
            operacao=Auditoria.Operacao.ALTERACAO,
            campo="nome",
        )
        self.assertEqual(reg.valor_anterior, "Zé")
        self.assertEqual(reg.valor_novo, "José")

    def test_auditoria_e_imutavel(self):
        p = Pessoa.objects.create(nome="Zé", cpf="997", data_nascimento=nasc(30))
        reg = Auditoria.objects.filter(registro_id=str(p.pk)).first()
        with self.assertRaises(RuntimeError):
            reg.justificativa = "x"
            reg.save()


class BloqueioTest(TestCase):
    def test_inscricao_bloqueada_nao_edita(self):
        req = Pessoa.objects.create(nome="A", cpf="1", data_nascimento=nasc(40))
        insc = Inscricao.objects.create(requerente=req)
        insc.bloqueada = True
        insc.data_finalizacao = timezone.now()
        insc.save()  # a transição para bloqueada é permitida

        insc.telefone = "99999-9999"
        with self.assertRaises(ValidationError):
            insc.save()

    def test_alteracao_autorizada_passa(self):
        req = Pessoa.objects.create(nome="B", cpf="2", data_nascimento=nasc(40))
        insc = Inscricao.objects.create(requerente=req, bloqueada=True)
        insc.telefone = "1234"
        insc._alteracao_autorizada = True
        insc.save()  # não deve levantar
        insc.refresh_from_db()
        self.assertEqual(insc.telefone, "1234")


class ClassificacaoTest(TestCase):
    def _apta(self, cpf, filhos_ate_12):
        req = Pessoa.objects.create(nome=f"R{cpf}", cpf=cpf, data_nascimento=nasc(40))
        insc = Inscricao.objects.create(
            requerente=req,
            data_referencia=REF,
            habitacao_precaria_ou_risco=True,
            matricula_comprovada=True,
            status=Inscricao.Status.APTO,
        )
        m = MembroNucleo.objects.create(inscricao=insc, pessoa=req, parentesco="REQUERENTE")
        Renda.objects.create(membro=m, tipo="FORMAL", valor=Decimal("3000"))
        for i in range(filhos_ate_12):
            f = Pessoa.objects.create(
                nome=f"F{cpf}{i}", cpf=f"{cpf}f{i}", data_nascimento=nasc(5)
            )
            MembroNucleo.objects.create(inscricao=insc, pessoa=f, parentesco="FILHO")
        services.calcular_e_salvar(insc)
        return insc

    def test_desempate_por_filhos(self):
        a = self._apta("10", filhos_ate_12=2)
        b = self._apta("20", filhos_ate_12=1)
        services.classificar_todos()
        a.refresh_from_db(); b.refresh_from_db()
        # Mesma pontuação; A tem mais filhos <=12 → posição melhor (menor número).
        self.assertLess(a.classificacao.posicao, b.classificacao.posicao)

    def test_empate_marcado_para_sorteio(self):
        a = self._apta("30", filhos_ate_12=1)
        b = self._apta("40", filhos_ate_12=1)
        services.classificar_todos()
        a.refresh_from_db(); b.refresh_from_db()
        self.assertTrue(a.classificacao.empate_pendente_sorteio)
        self.assertTrue(b.classificacao.empate_pendente_sorteio)


class RequisitosTest(TestCase):
    def _inscricao(self, idade=40, renda="3000", brasileiro=True):
        req = Pessoa.objects.create(
            nome="R", cpf="req", data_nascimento=nasc(idade), sexo="M", brasileiro=brasileiro
        )
        insc = Inscricao.objects.create(requerente=req, data_referencia=REF)
        m = MembroNucleo.objects.create(inscricao=insc, pessoa=req, parentesco="REQUERENTE")
        Renda.objects.create(membro=m, tipo="FORMAL", valor=Decimal(renda))
        return insc

    def test_apto_com_flags_documentais(self):
        insc = self._inscricao()
        insc.residencia_5anos_comprovada = True
        insc.nao_proprietario_declarado = True
        insc.nao_beneficiado_declarado = True
        insc._alteracao_autorizada = True
        insc.save()
        itens = requisitos.avaliar(insc)
        self.assertTrue(requisitos.apto(itens))

    def test_inapto_menor_de_idade(self):
        insc = self._inscricao(idade=17)
        itens = requisitos.avaliar(insc)
        self.assertFalse(requisitos.apto(itens))
        r1 = next(i for i in itens if i.codigo == "R1")
        self.assertFalse(r1.ok)

    def test_inapto_renda_acima_do_teto(self):
        insc = self._inscricao(renda="9000")
        itens = requisitos.avaliar(insc)
        r5 = next(i for i in itens if i.codigo == "R5")
        self.assertFalse(r5.ok)


def _com_perfil(username, *grupos):
    from django.contrib.auth.models import Group

    u = User.objects.create_user(username, password="x")
    for g in grupos:
        u.groups.add(Group.objects.get(name=g))
    return u


class ViewsSmokeTest(TestCase):
    def setUp(self):
        self.user = _com_perfil("srv", "Atendente")
        self.client.force_login(self.user)

    def test_paginas_principais_respondem(self):
        for nome in ["cadastro:dashboard", "cadastro:inscricao_list",
                     "cadastro:inscricao_nova", "cadastro:classificacao"]:
            self.assertEqual(self.client.get(reverse(nome)).status_code, 200)

    def test_logout_via_post(self):
        # O logout do Django aceita apenas POST; o botão "sair" usa formulário POST.
        resp = self.client.post(reverse("logout"))
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(self.client.get(reverse("cadastro:dashboard")).status_code, 302)

    def test_login_obrigatorio(self):
        self.client.logout()
        resp = self.client.get(reverse("cadastro:dashboard"))
        self.assertEqual(resp.status_code, 302)
        self.assertIn("/entrar/", resp.url)

    def test_criar_inscricao_e_finalizar(self):
        cpf = gera_cpf("111444777")
        resp = self.client.post(
            reverse("cadastro:inscricao_nova"),
            {
                "nome": "Fulano", "cpf": cpf, "data_nascimento": "1986-01-01",
                "sexo": "M", "brasileiro": "on",
            },
        )
        self.assertEqual(resp.status_code, 302)
        insc = Inscricao.objects.get(requerente__cpf=cpf)

        # detalhe responde
        self.assertEqual(
            self.client.get(reverse("cadastro:inscricao_detalhe", args=[insc.pk])).status_code, 200
        )
        # finalizar bloqueia
        self.client.post(reverse("cadastro:finalizar", args=[insc.pk]))
        insc.refresh_from_db()
        self.assertTrue(insc.bloqueada)
        self.assertIsNotNone(insc.data_finalizacao)


class RelatoriosTest(TestCase):
    def setUp(self):
        self.user = User.objects.create_user("srv2", password="x")
        self.client.force_login(self.user)
        req = Pessoa.objects.create(nome="Rel", cpf="rel1", data_nascimento=nasc(40), sexo="M")
        self.insc = Inscricao.objects.create(
            requerente=req, data_referencia=REF, habitacao_precaria_ou_risco=True,
            matricula_comprovada=True, status=Inscricao.Status.APTO,
        )
        m = MembroNucleo.objects.create(inscricao=self.insc, pessoa=req, parentesco="REQUERENTE")
        Renda.objects.create(membro=m, tipo="FORMAL", valor=Decimal("3000"))
        f = Pessoa.objects.create(nome="Fi", cpf="rel2", data_nascimento=nasc(5), sexo="F")
        MembroNucleo.objects.create(inscricao=self.insc, pessoa=f, parentesco="FILHO")
        services.calcular_e_salvar(self.insc)
        services.classificar_todos()

    def test_indice_relatorios(self):
        self.assertEqual(self.client.get(reverse("cadastro:relatorios")).status_code, 200)

    def test_exportacoes_xlsx(self):
        XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        for nome in ["rel_base", "rel_classificacao_xlsx", "rel_pendentes",
                     "rel_indeferidos", "rel_aptos", "rel_empates", "rel_auditoria"]:
            resp = self.client.get(reverse(f"cadastro:{nome}"))
            self.assertEqual(resp.status_code, 200, nome)
            self.assertEqual(resp["Content-Type"], XLSX, nome)
            self.assertTrue(resp.content[:2] == b"PK", f"{nome} não parece um .xlsx")

    def test_pdfs(self):
        for url in [reverse("cadastro:rel_classificacao_pdf"),
                    reverse("cadastro:rel_ficha_pdf", args=[self.insc.pk])]:
            resp = self.client.get(url)
            self.assertEqual(resp.status_code, 200)
            self.assertEqual(resp["Content-Type"], "application/pdf")
            self.assertTrue(resp.content[:5] == b"%PDF-", "não parece um PDF")

    def test_filtro_base_pcd(self):
        # Sem PcD no núcleo → filtro pcd deve retornar planilha sem linhas de dados.
        resp = self.client.get(reverse("cadastro:rel_base"), {"pcd": "1"})
        self.assertEqual(resp.status_code, 200)


class AcessoFluxoTest(TestCase):
    def _inscricao_apta(self, cpf="fx1"):
        req = Pessoa.objects.create(nome="Fx", cpf=cpf, data_nascimento=nasc(40), sexo="M")
        insc = Inscricao.objects.create(
            requerente=req, data_referencia=REF, status=Inscricao.Status.RECEBIDA,
            habitacao_precaria_ou_risco=True, matricula_comprovada=True,
            residencia_5anos_comprovada=True, nao_proprietario_declarado=True,
            nao_beneficiado_declarado=True, bloqueada=True,
        )
        m = MembroNucleo.objects.create(inscricao=insc, pessoa=req, parentesco="REQUERENTE")
        Renda.objects.create(membro=m, tipo="FORMAL", valor=Decimal("3000"))
        f = Pessoa.objects.create(nome="Fi", cpf=cpf + "f", data_nascimento=nasc(5))
        MembroNucleo.objects.create(inscricao=insc, pessoa=f, parentesco="FILHO")
        services.calcular_e_salvar(insc)
        return insc

    def test_consulta_nao_cria_inscricao(self):
        self.client.force_login(_com_perfil("cons", "Consulta"))
        resp = self.client.get(reverse("cadastro:inscricao_nova"))
        self.assertEqual(resp.status_code, 403)

    def test_atendente_nao_gera_classificacao(self):
        self.client.force_login(_com_perfil("at", "Atendente"))
        resp = self.client.post(reverse("cadastro:classificacao"))
        self.assertEqual(resp.status_code, 403)

    def test_analista_inicia_e_valida_analise(self):
        insc = self._inscricao_apta("fx2")
        self.client.force_login(_com_perfil("an", "Analista"))
        self.client.post(reverse("cadastro:transicionar", args=[insc.pk]),
                         {"destino": Inscricao.Status.EM_ANALISE})
        insc.refresh_from_db(); self.assertEqual(insc.status, Inscricao.Status.EM_ANALISE)
        self.client.post(reverse("cadastro:transicionar", args=[insc.pk]),
                         {"destino": Inscricao.Status.DOC_VALIDADA})
        insc.refresh_from_db(); self.assertEqual(insc.status, Inscricao.Status.DOC_VALIDADA)

    def test_analista_nao_homologa(self):
        insc = self._inscricao_apta("fx3")
        insc.status = Inscricao.Status.APTO
        insc._alteracao_autorizada = True
        insc.save()
        self.client.force_login(_com_perfil("an2", "Analista"))
        self.client.post(reverse("cadastro:transicionar", args=[insc.pk]),
                         {"destino": Inscricao.Status.HOMOLOGADO})
        insc.refresh_from_db()
        self.assertEqual(insc.status, Inscricao.Status.APTO)  # não mudou

    def test_comissao_homologa(self):
        insc = self._inscricao_apta("fx4")
        insc.status = Inscricao.Status.APTO
        insc._alteracao_autorizada = True
        insc.save()
        self.client.force_login(_com_perfil("com", "Comissao"))
        self.client.post(reverse("cadastro:transicionar", args=[insc.pk]),
                         {"destino": Inscricao.Status.HOMOLOGADO})
        insc.refresh_from_db()
        self.assertEqual(insc.status, Inscricao.Status.HOMOLOGADO)

    def test_apto_bloqueado_se_requisito_pendente(self):
        # Núcleo sem os flags documentais → não apto → transição para APTO barrada.
        req = Pessoa.objects.create(nome="Z", cpf="fx5", data_nascimento=nasc(40))
        insc = Inscricao.objects.create(
            requerente=req, data_referencia=REF, status=Inscricao.Status.DOC_VALIDADA
        )
        m = MembroNucleo.objects.create(inscricao=insc, pessoa=req, parentesco="REQUERENTE")
        Renda.objects.create(membro=m, tipo="FORMAL", valor=Decimal("3000"))
        self.client.force_login(_com_perfil("an3", "Analista"))
        self.client.post(reverse("cadastro:transicionar", args=[insc.pk]),
                         {"destino": Inscricao.Status.APTO})
        insc.refresh_from_db()
        self.assertEqual(insc.status, Inscricao.Status.DOC_VALIDADA)  # barrado


class InjecaoPlanilhaTest(TestCase):
    def test_neutraliza_formula(self):
        from io import BytesIO
        from openpyxl import load_workbook
        from cadastro.relatorios import planilha_response

        resp = planilha_response(
            "t.xlsx", ["Nome"], [["=HYPERLINK(\"http://x\")"], ["+1"], ["João"]]
        )
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        # Linha 1 é cabeçalho; dados começam na linha 2.
        self.assertEqual(ws.cell(row=2, column=1).value, "'=HYPERLINK(\"http://x\")")
        self.assertEqual(ws.cell(row=3, column=1).value, "'+1")
        self.assertEqual(ws.cell(row=4, column=1).value, "João")
        # Nenhuma célula de dado é do tipo fórmula.
        for r in (2, 3, 4):
            self.assertNotEqual(ws.cell(row=r, column=1).data_type, "f")


class WizardTest(TestCase):
    def setUp(self):
        self.user = _com_perfil("at_wiz", "Atendente")
        self.client.force_login(self.user)
        self.cpf = gera_cpf("222333444")
        resp = self.client.post(reverse("cadastro:inscricao_nova"), {
            "nome": "Assistente Teste", "cpf": self.cpf,
            "data_nascimento": "1986-01-01", "sexo": "M", "brasileiro": "on",
        })
        self.insc = Inscricao.objects.get(requerente__cpf=self.cpf)

    def test_nova_entra_no_wizard(self):
        resp = self.client.post(reverse("cadastro:inscricao_nova"), {
            "nome": "Outro", "cpf": gera_cpf("333444555"),
            "data_nascimento": "1986-01-01", "sexo": "F", "brasileiro": "on",
        })
        self.assertIn("/cadastro/requerente/", resp.url)

    def test_data_nascimento_aparece_no_assistente(self):
        # A data informada na criação deve vir preenchida (value ISO) no assistente.
        url = reverse("cadastro:wizard", args=[self.insc.pk, "requerente"])
        html = self.client.get(url).content.decode()
        self.assertIn('value="1986-01-01"', html)

    def test_todas_etapas_respondem(self):
        for etapa in ["requerente", "nucleo", "renda", "documentos", "avaliacao", "revisao"]:
            url = reverse("cadastro:wizard", args=[self.insc.pk, etapa])
            self.assertEqual(self.client.get(url).status_code, 200, etapa)

    def test_salvar_e_avancar(self):
        url = reverse("cadastro:wizard", args=[self.insc.pk, "requerente"])
        resp = self.client.post(url, {
            "nome": "Assistente Teste", "cpf": self.cpf, "data_nascimento": "1986-01-01",
            "sexo": "M", "estado_civil": "SOLTEIRO", "brasileiro": "on", "acao": "avancar",
        })
        self.assertEqual(resp.status_code, 302)
        self.assertIn("/cadastro/nucleo/", resp.url)

    def test_revisao_lista_pendencias(self):
        from cadastro.wizard import pendencias
        pend = pendencias(self.insc)
        # Recém-criada: deve haver pendências (contato, núcleo, documentos...).
        self.assertTrue(pend)
        resp = self.client.get(reverse("cadastro:wizard", args=[self.insc.pk, "revisao"]))
        self.assertContains(resp, "Pendências")

    def test_etapa_invalida_404(self):
        resp = self.client.get(reverse("cadastro:wizard", args=[self.insc.pk, "inexistente"]))
        self.assertEqual(resp.status_code, 404)


class ExclusaoTest(TestCase):
    def _inscricao(self, cpf="ex1"):
        req = Pessoa.objects.create(nome="Excluir", cpf=cpf, data_nascimento=nasc(40))
        insc = Inscricao.objects.create(requerente=req)
        MembroNucleo.objects.create(inscricao=insc, pessoa=req, parentesco="REQUERENTE")
        return insc

    def test_admin_exclui(self):
        from django.contrib.auth.models import Group
        insc = self._inscricao("ex_admin")
        u = User.objects.create_user("adm", password="x")
        u.groups.add(Group.objects.get(name="Administrador"))
        self.client.force_login(u)
        resp = self.client.post(reverse("cadastro:inscricao_excluir", args=[insc.pk]))
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(Inscricao.objects.filter(pk=insc.pk).exists())

    def test_exclusao_libera_cpf(self):
        # Após excluir, a pessoa órfã é removida e o CPF fica livre para novo cadastro.
        from django.contrib.auth.models import Group
        insc = self._inscricao("07065903680")
        u = User.objects.create_user("adm3", password="x")
        u.groups.add(Group.objects.get(name="Administrador"))
        self.client.force_login(u)
        self.client.post(reverse("cadastro:inscricao_excluir", args=[insc.pk]))
        self.assertFalse(Pessoa.objects.filter(cpf="07065903680").exists())
        # Recriar com o mesmo CPF deve funcionar (sem erro de duplicidade).
        resp = self.client.post(reverse("cadastro:inscricao_nova"), {
            "nome": "Novo Cadastro", "cpf": "07065903680",
            "data_nascimento": "1984-08-08", "sexo": "M", "brasileiro": "on",
        })
        self.assertEqual(resp.status_code, 302)
        self.assertTrue(Pessoa.objects.filter(cpf="07065903680").exists())

    def test_pessoa_compartilhada_nao_e_removida(self):
        # Pessoa que também é membro de OUTRO núcleo não deve ser apagada.
        from django.contrib.auth.models import Group
        compartilhada = Pessoa.objects.create(nome="Comum", cpf="comum1", data_nascimento=nasc(30))
        insc1 = self._inscricao("dono1")
        MembroNucleo.objects.create(inscricao=insc1, pessoa=compartilhada, parentesco="OUTRO")
        outra = self._inscricao("dono2")
        MembroNucleo.objects.create(inscricao=outra, pessoa=compartilhada, parentesco="OUTRO")
        u = User.objects.create_user("adm4", password="x")
        u.groups.add(Group.objects.get(name="Administrador"))
        self.client.force_login(u)
        self.client.post(reverse("cadastro:inscricao_excluir", args=[insc1.pk]))
        self.assertTrue(Pessoa.objects.filter(cpf="comum1").exists())  # ainda em 'outra'

    def test_atendente_nao_exclui(self):
        insc = self._inscricao("ex_at")
        self.client.force_login(_com_perfil("at_del", "Atendente"))
        resp = self.client.post(reverse("cadastro:inscricao_excluir", args=[insc.pk]))
        self.assertEqual(resp.status_code, 403)
        self.assertTrue(Inscricao.objects.filter(pk=insc.pk).exists())

    def test_botao_excluir_so_para_admin_na_lista(self):
        from django.contrib.auth.models import Group
        insc = self._inscricao("ex_lista")
        url_excluir = reverse("cadastro:inscricao_excluir", args=[insc.pk])
        # Atendente não vê o formulário de exclusão
        self.client.force_login(_com_perfil("at_lista", "Atendente"))
        html = self.client.get(reverse("cadastro:inscricao_list")).content.decode()
        self.assertNotIn(url_excluir, html)
        # Admin vê o formulário de exclusão
        u = User.objects.create_user("adm2", password="x")
        u.groups.add(Group.objects.get(name="Administrador"))
        self.client.force_login(u)
        html = self.client.get(reverse("cadastro:inscricao_list")).content.decode()
        self.assertIn(url_excluir, html)


class ReciboTest(TestCase):
    def setUp(self):
        self.user = _com_perfil("at_rec", "Atendente")
        self.client.force_login(self.user)
        self.req = Pessoa.objects.create(nome="Recibo", cpf="55501", data_nascimento=nasc(40))
        self.insc = Inscricao.objects.create(requerente=self.req)
        MembroNucleo.objects.create(inscricao=self.insc, pessoa=self.req, parentesco="REQUERENTE")

    def test_recibo_indisponivel_antes_de_finalizar(self):
        resp = self.client.get(reverse("cadastro:rel_recibo_pdf", args=[self.insc.pk]))
        self.assertEqual(resp.status_code, 302)  # redireciona com aviso

    def test_finalizar_gera_protocolo_e_recibo(self):
        self.client.post(reverse("cadastro:finalizar", args=[self.insc.pk]))
        self.insc.refresh_from_db()
        self.assertTrue(self.insc.protocolo)
        self.assertTrue(self.insc.protocolo.startswith("MCMV-"))
        self.assertIn(self.insc.numero_inscricao, self.insc.protocolo)
        resp = self.client.get(reverse("cadastro:rel_recibo_pdf", args=[self.insc.pk]))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], "application/pdf")
        self.assertTrue(resp.content[:5] == b"%PDF-")


class CpfValidacaoTest(TestCase):
    def setUp(self):
        self.client.force_login(_com_perfil("at_cpf", "Atendente"))

    def test_cpf_invalido_rejeitado(self):
        resp = self.client.post(reverse("cadastro:inscricao_nova"), {
            "nome": "Teste", "cpf": "12345678900",  # dígitos verificadores errados
            "data_nascimento": "1986-01-01", "sexo": "M", "brasileiro": "on",
        })
        self.assertEqual(resp.status_code, 200)  # reexibe o form
        self.assertContains(resp, "CPF inválido")
        self.assertFalse(Inscricao.objects.filter(requerente__cpf="12345678900").exists())

    def test_cpf_valido_aceito_e_normalizado(self):
        cpf = gera_cpf("529982247")  # gera dígitos válidos
        formatado = f"{cpf[:3]}.{cpf[3:6]}.{cpf[6:9]}-{cpf[9:]}"
        resp = self.client.post(reverse("cadastro:inscricao_nova"), {
            "nome": "Teste", "cpf": formatado,  # com pontuação
            "data_nascimento": "1986-01-01", "sexo": "M", "brasileiro": "on",
        })
        self.assertEqual(resp.status_code, 302)
        # Deve ter sido salvo só com dígitos.
        self.assertTrue(Inscricao.objects.filter(requerente__cpf=cpf).exists())

    def test_algoritmo_cpf(self):
        from cadastro.validadores import cpf_valido
        self.assertTrue(cpf_valido("070.659.036-80"))
        self.assertFalse(cpf_valido("111.111.111-11"))
        self.assertFalse(cpf_valido("123"))
