"""Testes dos relatórios (agendamentos, BPA, indicadores e exportação)."""
from datetime import date, time, timedelta

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse

from apps.agendamentos.models import Agendamento, StatusAgendamento
from apps.core.models import Municipio
from apps.destinos.models import Destino
from apps.pacientes.models import Paciente

from .filtros import filtrar

User = get_user_model()


def proxima_sexta(base=None):
    d = base or date.today()
    while d.weekday() != 4:
        d += timedelta(days=1)
    return d


class BaseDados(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="c", password="x", perfil="CONSULTA")
        self.mun = Municipio.objects.create(codigo_ibge="3131307", nome="Itabira")
        self.pac = Paciente.objects.create(
            nome="Maria", cpf="52998224725", data_nascimento=date(1970, 1, 1),
            sexo="F", telefone_principal="31999990000", municipio=self.mun, raca_cor="03",
        )
        self.dst = Destino.objects.create(nome="Hospital X", municipio=self.mun)
        self.dia = proxima_sexta()
        self.ag = Agendamento.objects.create(
            paciente=self.pac, destino=self.dst, data=self.dia, horario=time(7, 0),
            procedimento="Cardiologia", status=StatusAgendamento.FINALIZADO,
        )
        Agendamento.objects.create(
            paciente=self.pac, destino=self.dst, data=self.dia, horario=time(8, 0),
            status=StatusAgendamento.CANCELADO,
        )


class FiltroTest(BaseDados):
    def test_filtra_por_dia(self):
        qs, resumo = filtrar({"dia": self.dia.strftime("%Y-%m-%d")}, incluir_cancelados=True)
        self.assertEqual(qs.count(), 2)

    def test_exclui_cancelados_por_padrao(self):
        qs, _ = filtrar({}, incluir_cancelados=False)
        self.assertEqual(qs.count(), 1)

    def test_filtra_por_nome_e_procedimento(self):
        qs, _ = filtrar({"q": "mar", "procedimento": "cardio"}, incluir_cancelados=True)
        self.assertEqual(qs.count(), 1)


class ViewsTest(BaseDados):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.user)

    def test_paginas_carregam(self):
        for nome in ["index", "agendamentos", "bpa", "indicadores"]:
            self.assertEqual(self.client.get(reverse(f"relatorios:{nome}")).status_code, 200)

    def test_exige_login(self):
        self.client.logout()
        self.assertEqual(self.client.get(reverse("relatorios:agendamentos")).status_code, 302)

    def test_export_xlsx(self):
        r = self.client.get(reverse("relatorios:agendamentos"), {"export": "xlsx"})
        self.assertEqual(r.status_code, 200)
        self.assertIn("spreadsheet", r["Content-Type"])
        self.assertEqual(r.content[:2], b"PK")  # xlsx é um zip

    def test_export_pdf_bpa(self):
        r = self.client.get(reverse("relatorios:bpa"), {"export": "pdf", "mes": self.dia.strftime("%Y-%m")})
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r["Content-Type"], "application/pdf")
        self.assertTrue(r.content.startswith(b"%PDF"))


class MapaViagemTest(TestCase):
    """Mapa de Viagem: agrupamento por veículo e exportação."""

    def setUp(self):
        self.user = User.objects.create_user(username="c", password="x", perfil="CONSULTA")
        self.mun = Municipio.objects.create(codigo_ibge="3106200", nome="Belo Horizonte")
        self.pac = Paciente.objects.create(
            nome="José Silva", cpf="52998224725", data_nascimento=date(1960, 1, 1),
            sexo="M", telefone_principal="31988887777", municipio=self.mun, raca_cor="01",
        )
        self.pac2 = Paciente.objects.create(
            nome="Ana Souza", cns="700000000000000", data_nascimento=date(1975, 2, 2),
            sexo="F", telefone_principal="31977776666", municipio=self.mun, raca_cor="03",
        )
        self.dst = Destino.objects.create(
            nome="Hospital das Clínicas", municipio=self.mun, endereco="Av. Alfredo Balena 190",
        )
        self.dia = proxima_sexta()
        # CARRO 10 e CARRO 2 para checar ordenação natural (2 antes de 10).
        Agendamento.objects.create(
            paciente=self.pac2, destino=self.dst, data=self.dia, horario=time(9, 0),
            tipo_veiculo="CARRO 10", local_embarque="Rua A 100", hora_embarque=time(6, 30),
        )
        self.ag = Agendamento.objects.create(
            paciente=self.pac, destino=self.dst, data=self.dia, horario=time(8, 0),
            tipo_veiculo="CARRO 2", local_embarque="Rua B 200", hora_embarque=time(5, 0),
            acompanhante="Maria Acompanhante",
        )
        # Sem veículo → grupo "A DEFINIR", ao final.
        Agendamento.objects.create(
            paciente=self.pac, destino=self.dst, data=self.dia, horario=time(10, 0),
        )
        # Cancelado não entra no mapa.
        Agendamento.objects.create(
            paciente=self.pac, destino=self.dst, data=self.dia, horario=time(11, 0),
            tipo_veiculo="CARRO 2", status=StatusAgendamento.CANCELADO,
        )

    def _grupos(self):
        from .mapa import agrupar_por_veiculo
        qs = (
            Agendamento.objects.filter(data=self.dia)
            .exclude(status=StatusAgendamento.CANCELADO)
            .select_related("paciente", "destino", "destino__municipio")
        )
        return agrupar_por_veiculo(qs)

    def test_agrupamento_ordem_natural_e_sem_veiculo_ao_final(self):
        grupos = self._grupos()
        nomes = [g["veiculo"] for g in grupos]
        self.assertEqual(nomes, ["CARRO 2", "CARRO 10", "A DEFINIR"])

    def test_acompanhante_vira_linha_e_saida_e_a_mais_cedo(self):
        grupos = {g["veiculo"]: g for g in self._grupos()}
        carro2 = grupos["CARRO 2"]
        # 1 paciente + 1 acompanhante = 2 linhas.
        self.assertEqual(len(carro2["linhas"]), 2)
        self.assertTrue(carro2["linhas"][1]["ac"])
        self.assertEqual(carro2["linhas"][1]["nome"], "AC. Maria Acompanhante")
        self.assertEqual(carro2["saida"], time(5, 0))

    def test_cancelado_nao_entra(self):
        total = sum(1 for g in self._grupos() for lin in g["linhas"] if not lin["ac"])
        self.assertEqual(total, 3)  # 4 agendamentos, 1 cancelado

    def test_local_inclui_endereco(self):
        grupos = {g["veiculo"]: g for g in self._grupos()}
        self.assertIn("Av. Alfredo Balena 190", grupos["CARRO 10"]["linhas"][0]["local"])

    def test_pagina_carrega(self):
        self.client.force_login(self.user)
        r = self.client.get(reverse("relatorios:mapa_viagem"), {"data": self.dia.strftime("%Y-%m-%d")})
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, "CARRO 2")
        self.assertContains(r, "AC. Maria Acompanhante")

    def test_export_xlsx(self):
        self.client.force_login(self.user)
        r = self.client.get(
            reverse("relatorios:mapa_viagem"),
            {"data": self.dia.strftime("%Y-%m-%d"), "export": "xlsx", "motorista": "João"},
        )
        self.assertEqual(r.status_code, 200)
        self.assertIn("spreadsheet", r["Content-Type"])
        self.assertEqual(r.content[:2], b"PK")

    def test_export_pdf(self):
        self.client.force_login(self.user)
        r = self.client.get(
            reverse("relatorios:mapa_viagem"),
            {"data": self.dia.strftime("%Y-%m-%d"), "export": "pdf"},
        )
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r["Content-Type"], "application/pdf")
        self.assertTrue(r.content.startswith(b"%PDF"))

    def test_dia_sem_agendamentos_nao_quebra(self):
        self.client.force_login(self.user)
        vazio = (self.dia + timedelta(days=7)).strftime("%Y-%m-%d")
        for exp in ["xlsx", "pdf"]:
            r = self.client.get(reverse("relatorios:mapa_viagem"), {"data": vazio, "export": exp})
            self.assertEqual(r.status_code, 200)


class FormulaInjectionTest(TestCase):
    def test_sanitiza_gatilhos_de_formula(self):
        from openpyxl import load_workbook
        from io import BytesIO

        from .exports import _sanitizar, exportar_xlsx

        # Campo de texto malicioso que iniciaria uma fórmula no Excel.
        self.assertEqual(_sanitizar("=HYPERLINK(1)"), "'=HYPERLINK(1)")
        self.assertEqual(_sanitizar("+1+1"), "'+1+1")
        self.assertEqual(_sanitizar("texto normal"), "texto normal")

        conteudo = exportar_xlsx("T", ["A"], [["=1+1"], ["ok"]])
        wb = load_workbook(BytesIO(conteudo))
        ws = wb.active
        # A célula é armazenada como texto neutralizado, não como fórmula.
        self.assertEqual(ws["A2"].value, "'=1+1")
        self.assertEqual(ws["A2"].data_type, "s")
