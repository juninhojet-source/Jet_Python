"""Exportação de relatórios em Excel (XLSX) e PDF."""
from io import BytesIO

from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

NAVY = "16386E"


def exportar_xlsx(titulo, colunas, linhas):
    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]
    fill = PatternFill("solid", fgColor=NAVY)
    fonte = Font(bold=True, color="FFFFFF")
    for c, nome in enumerate(colunas, start=1):
        cel = ws.cell(row=1, column=c, value=nome)
        cel.fill = fill
        cel.font = fonte
        cel.alignment = Alignment(horizontal="left", vertical="center")
    for r, linha in enumerate(linhas, start=2):
        for c, valor in enumerate(linha, start=1):
            ws.cell(row=r, column=c, value=valor)
    # largura das colunas
    for c, nome in enumerate(colunas, start=1):
        largura = max(len(str(nome)), *(len(str(l[c - 1])) for l in linhas)) if linhas else len(str(nome))
        ws.column_dimensions[get_column_letter(c)].width = min(max(largura + 2, 10), 45)
    ws.freeze_panes = "A2"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def exportar_pdf(titulo, subtitulo, colunas, linhas):
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=14 * mm, rightMargin=14 * mm, topMargin=14 * mm, bottomMargin=14 * mm,
        title=titulo,
    )
    estilos = getSampleStyleSheet()
    h = estilos["Title"]; h.textColor = colors.HexColor("#" + NAVY); h.fontSize = 16
    sub = estilos["Normal"]; sub.fontSize = 9; sub.textColor = colors.HexColor("#5b6b7b")
    cel = estilos["Normal"].clone("cel"); cel.fontSize = 7.5; cel.leading = 9

    elementos = [
        Paragraph(f"{settings.ORGAO['SISTEMA_NOME']} — {titulo}", h),
        Paragraph(f"{settings.ORGAO['SECRETARIA']} · {subtitulo}", sub),
        Spacer(1, 8),
    ]
    dados = [[Paragraph(f"<b>{c}</b>", cel) for c in colunas]]
    for linha in linhas:
        dados.append([Paragraph(str(v) if v is not None else "", cel) for v in linha])

    tabela = Table(dados, repeatRows=1)
    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + NAVY)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#c9d3e0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f6fc")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    elementos.append(tabela)
    doc.build(elementos)
    return buf.getvalue()
