"""Mapa de Viagem (Garagem).

Gera o mapa diário de viagens no mesmo formato usado pela Garagem: os
agendamentos do dia agrupados por veículo, com embarque, destino, local e
horário de saída, além da legenda de conferência (N/P/D/R).

O agrupamento usa o campo "Tipo de veículo" do agendamento (preenchimento
livre — ex.: "CARRO 1", "AMBULÂNCIA 2", "MICRO ITABIRA"). Agendamentos sem
veículo definido ficam no grupo "A DEFINIR", ao final.
"""
import re
from io import BytesIO

from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

NAVY = "16386E"
GOLD = "F2C230"
CINZA = "F2F6FC"

SEM_VEICULO = "A DEFINIR"

LEGENDA = [
    "N = não compareceu",
    "P = presente",
    "D = desceu do veículo",
    "R = retornou ao veículo",
]

COLUNAS = [
    "Veículo", "Nome do paciente", "Telefone", "Horário",
    "Embarque", "Destino", "Local", "Horário saída",
]

# Caracteres que iniciam uma fórmula no Excel/LibreOffice.
_GATILHOS_FORMULA = ("=", "+", "-", "@", "\t", "\r")


def _sanitizar(valor):
    if isinstance(valor, str) and valor[:1] in _GATILHOS_FORMULA:
        return "'" + valor
    return valor


def _hhmm(t):
    return t.strftime("%H:%M") if t else ""


def _natural_key(texto):
    """Ordena de forma natural: CARRO 2 antes de CARRO 10."""
    return [int(p) if p.isdigit() else p.lower() for p in re.split(r"(\d+)", texto)]


def _local_destino(destino):
    partes = [destino.nome]
    if destino.endereco:
        partes.append(destino.endereco)
    return " — ".join(partes)


def agrupar_por_veiculo(qs):
    """Agrupa um queryset de agendamentos por veículo, pronto para render.

    Retorna uma lista de grupos, cada um com:
        {"veiculo": str, "saida": time|None, "linhas": [ {..paciente..}, {..AC..} ]}
    """
    grupos = {}
    for a in qs:
        chave = (a.tipo_veiculo or "").strip() or SEM_VEICULO
        grupo = grupos.setdefault(chave, {"veiculo": chave, "saida": None, "linhas": []})

        # Horário de saída da garagem = embarque mais cedo do grupo.
        if a.hora_embarque and (grupo["saida"] is None or a.hora_embarque < grupo["saida"]):
            grupo["saida"] = a.hora_embarque

        grupo["linhas"].append({
            "nome": a.paciente.nome,
            "telefone": a.contato or a.paciente.telefone_principal or "",
            "horario": _hhmm(a.horario),
            "embarque": a.local_embarque or "",
            "destino": a.destino.municipio.nome if a.destino.municipio_id else "",
            "local": _local_destino(a.destino),
            "ac": False,
        })
        if a.acompanhante.strip():
            grupo["linhas"].append({
                "nome": f"AC. {a.acompanhante.strip()}",
                "telefone": "", "horario": "", "embarque": "",
                "destino": "", "local": "", "ac": True,
            })

    reais = sorted(
        (g for k, g in grupos.items() if k != SEM_VEICULO),
        key=lambda g: _natural_key(g["veiculo"]),
    )
    if SEM_VEICULO in grupos:
        reais.append(grupos[SEM_VEICULO])
    return reais


# --------------------------------------------------------------------- Excel
def gerar_mapa_xlsx(dia, grupos, motorista=""):
    wb = Workbook()
    ws = wb.active
    ws.title = "Mapa de Viagem"

    n_cols = len(COLUNAS)
    ult = get_column_letter(n_cols)
    fina = Side(style="thin", color="B7C3D6")
    borda = Border(left=fina, right=fina, top=fina, bottom=fina)
    centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    esq = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def fundir(c1, c2):
        ws.merge_cells(f"{c1}:{c2}")

    # Cabeçalho institucional
    fundir(f"A1", f"{ult}1")
    ws["A1"] = settings.ORGAO["SECRETARIA"] + " de " + "Barão de Cocais"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = centro
    ws.row_dimensions[1].height = 26

    fundir(f"A2", f"{ult}2")
    ws["A2"] = "Agendamentos de transporte — GARAGEM"
    ws["A2"].font = Font(name="Arial", bold=True, size=11, color=NAVY)
    ws["A2"].fill = PatternFill("solid", fgColor=GOLD)
    ws["A2"].alignment = centro
    ws.row_dimensions[2].height = 20

    # Linha DATA / MOTORISTA
    ws["A4"] = "DATA:"
    ws["A4"].font = Font(name="Arial", bold=True)
    ws["B4"] = dia.strftime("%d/%m/%Y")
    ws["B4"].font = Font(name="Arial")
    ws["E4"] = "MOTORISTA:"
    ws["E4"].font = Font(name="Arial", bold=True)
    fundir("F4", ult + "4")
    ws["F4"] = motorista
    ws["F4"].font = Font(name="Arial")

    # Cabeçalho da tabela
    linha_cab = 6
    for c, nome in enumerate(COLUNAS, start=1):
        cel = ws.cell(row=linha_cab, column=c, value=nome)
        cel.font = Font(name="Arial", bold=True, color="FFFFFF")
        cel.fill = PatternFill("solid", fgColor=NAVY)
        cel.alignment = centro
        cel.border = borda
    ws.row_dimensions[linha_cab].height = 22

    # Dados agrupados por veículo
    r = linha_cab + 1
    for gi, grupo in enumerate(grupos):
        inicio = r
        for linha in grupo["linhas"]:
            valores = [
                "",  # A (veículo) — mesclado ao final do grupo
                linha["nome"], linha["telefone"], linha["horario"],
                linha["embarque"], linha["destino"], linha["local"],
                "",  # H (saída) — mesclado ao final do grupo
            ]
            for c, v in enumerate(valores, start=1):
                cel = ws.cell(row=r, column=c, value=_sanitizar(v))
                cel.border = borda
                cel.alignment = esq if c in (2, 5, 7) else centro
                fonte_kwargs = {"name": "Arial", "size": 10}
                if linha["ac"]:
                    fonte_kwargs.update(italic=True, color="5B6B7B")
                cel.font = Font(**fonte_kwargs)
            if gi % 2 == 1:
                for c in range(2, n_cols):  # não pinta A/H (mesclados)
                    ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=CINZA)
            r += 1
        fim = r - 1
        if fim < inicio:
            continue
        # Veículo (coluna A) e Saída (coluna H) mesclados no bloco do grupo
        if fim > inicio:
            fundir(f"A{inicio}", f"A{fim}")
            fundir(f"{ult}{inicio}", f"{ult}{fim}")
        ws[f"A{inicio}"] = grupo["veiculo"]
        ws[f"A{inicio}"].font = Font(name="Arial", bold=True, size=10, color=NAVY)
        ws[f"A{inicio}"].alignment = centro
        ws[f"{ult}{inicio}"] = _hhmm(grupo["saida"])
        ws[f"{ult}{inicio}"].font = Font(name="Arial", bold=True, size=11, color="B02A24")
        ws[f"{ult}{inicio}"].alignment = centro
        for rr in range(inicio, fim + 1):
            ws[f"A{rr}"].border = borda
            ws[f"{ult}{rr}"].border = borda

    if not grupos:
        fundir(f"A{r}", f"{ult}{r}")
        ws.cell(row=r, column=1, value="Nenhum agendamento para a data selecionada.").alignment = centro
        r += 1

    # Legenda
    r += 1
    ws.cell(row=r, column=1, value="Legenda:").font = Font(name="Arial", bold=True)
    for texto in LEGENDA:
        r += 1
        ws.cell(row=r, column=1, value=texto).font = Font(name="Arial", size=9, color="5B6B7B")

    # Larguras
    larguras = {1: 15, 2: 34, 3: 14, 4: 9, 5: 34, 6: 13, 7: 40, 8: 10}
    for c, w in larguras.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = f"A{linha_cab + 1}"

    # Impressão em paisagem, ajustado à largura da página
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f"1:{linha_cab}"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ----------------------------------------------------------------------- PDF
def gerar_mapa_pdf(dia, grupos, motorista=""):
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=10 * mm, rightMargin=10 * mm, topMargin=10 * mm, bottomMargin=10 * mm,
        title="Mapa de Viagem",
    )
    estilos = getSampleStyleSheet()
    titulo = ParagraphStyle(
        "titulo", parent=estilos["Title"], fontSize=14,
        textColor=colors.HexColor("#" + NAVY), spaceAfter=2,
    )
    sub = ParagraphStyle("sub", parent=estilos["Normal"], fontSize=9,
                         textColor=colors.HexColor("#5b6b7b"))
    cel = ParagraphStyle("cel", parent=estilos["Normal"], fontSize=7.5, leading=9)
    cel_ac = ParagraphStyle("celac", parent=cel, textColor=colors.HexColor("#5b6b7b"),
                            fontName="Helvetica-Oblique")

    elementos = [
        Paragraph(settings.ORGAO["SECRETARIA"] + " de Barão de Cocais", titulo),
        Paragraph("Agendamentos de transporte — GARAGEM", sub),
        Paragraph(f"<b>DATA:</b> {dia:%d/%m/%Y} &nbsp;&nbsp; <b>MOTORISTA:</b> {motorista or '________________'}", sub),
        Spacer(1, 8),
    ]

    dados = [[Paragraph(f"<b>{c}</b>", cel) for c in COLUNAS]]
    estilo_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + NAVY)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#b7c3d6")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]

    linha_atual = 1
    for grupo in grupos:
        inicio = linha_atual
        for lin in grupo["linhas"]:
            est = cel_ac if lin["ac"] else cel
            dados.append([
                Paragraph("", cel),  # veículo — preenchido via SPAN
                Paragraph(lin["nome"], est), Paragraph(lin["telefone"], est),
                Paragraph(lin["horario"], est), Paragraph(lin["embarque"], est),
                Paragraph(lin["destino"], est), Paragraph(lin["local"], est),
                Paragraph("", cel),  # saída — preenchido via SPAN
            ])
            linha_atual += 1
        fim = linha_atual - 1
        if fim < inicio:
            continue
        dados[inicio][0] = Paragraph(f"<b>{grupo['veiculo']}</b>", cel)
        dados[inicio][7] = Paragraph(f"<b>{_hhmm(grupo['saida'])}</b>", cel)
        estilo_cmds += [
            ("SPAN", (0, inicio), (0, fim)),
            ("SPAN", (7, inicio), (7, fim)),
            ("LINEABOVE", (0, inicio), (-1, inicio), 0.8, colors.HexColor("#" + NAVY)),
        ]

    if not grupos:
        dados.append([Paragraph("Nenhum agendamento para a data selecionada.", cel)] + [""] * (len(COLUNAS) - 1))
        estilo_cmds.append(("SPAN", (0, 1), (-1, 1)))

    larguras = [58, 150, 62, 40, 150, 58, 175, 45]
    tabela = Table(dados, repeatRows=1, colWidths=larguras)
    tabela.setStyle(TableStyle(estilo_cmds))
    elementos.append(tabela)

    elementos.append(Spacer(1, 8))
    elementos.append(Paragraph("<b>Legenda:</b> " + " · ".join(LEGENDA), sub))

    doc.build(elementos)
    return buf.getvalue()
